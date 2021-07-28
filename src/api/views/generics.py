from datetime import datetime
import os
import zipfile
from io import BytesIO

from django.utils import timezone

import xlsxwriter
from api import models, serializers
from django.contrib.auth.models import User
from django.db.models import F
from django.http import HttpResponse
from django.http.response import FileResponse, Http404
from django.shortcuts import get_object_or_404
from knox.auth import TokenAuthentication
from knox.crypto import hash_token
from knox.models import AuthToken
from knox.views import LoginView as KnoxLoginView
from rest_framework import generics, permissions, status, views
from rest_framework.authentication import BasicAuthentication, get_authorization_header
from rest_framework.response import Response
from rest_framework.parsers import FormParser, MultiPartParser
from openpyxl import load_workbook


def select_token_instance(request):
    auth = get_authorization_header(request).split()
    token = auth[1].decode()
    if token is not None:
        for auth_token in AuthToken.objects.filter(user=request.user):
            digest = hash_token(token, auth_token.salt)
            if digest == auth_token.digest:
                return token, auth_token
    return None, None


class LoginView(KnoxLoginView):
    authentication_classes = [BasicAuthentication, TokenAuthentication]

    def get(self, request, format=None):
        token, instance = select_token_instance(request)
        if token is not None:
            data = self.get_post_response_data(request, token, instance)
            return Response(data)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class ChangePasswordView(generics.UpdateAPIView):
    model = User
    permission_classes = [
        permissions.IsAuthenticated,
    ]
    serializer_class = serializers.ChangePasswordSerializer

    def get_object(self, queryset=None):
        obj = self.request.user
        return obj

    def update(self, request, *args, **kwargs):
        self.object = self.get_object()
        serializer = self.get_serializer(data=request.data)

        if serializer.is_valid():
            if not self.object.check_password(serializer.data.get("old_password")):
                return Response(
                    {"old_password": ["Senha incorreta."]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            self.object.set_password(serializer.data.get("new_password"))
            self.object.save()
            response = {
                "status": "success",
                "code": status.HTTP_200_OK,
                "message": "Senha atualizada com sucesso",
                "data": [],
            }
            return Response(response)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ChangeTipoDeSituacaoOrdemView(generics.RetrieveAPIView):
    model = User
    permission_classes = [
        permissions.IsAdminUser,
    ]
    serializer_class = serializers.Tipo_de_situacaoSerializer

    def get(self, request, *args, **kwargs):
        tipo_de_situacao_id = self.request.query_params.get("tipo_de_situacao_id", None)
        ordem_anterior = self.request.query_params.get("ordem_anterior", None)
        ordem_nova = self.request.query_params.get("ordem_nova", None)

        if tipo_de_situacao_id and ordem_anterior and ordem_nova:
            tipo_de_situacao = get_object_or_404(models.Tipo_de_situacao, id=tipo_de_situacao_id)
            tipo_de_situacao.ordem = 0
            tipo_de_situacao.save()

            if ordem_nova > ordem_anterior:
                tipos_de_situacoes = models.Tipo_de_situacao.objects.filter(
                    ordem__range=(ordem_anterior, ordem_nova)
                ).all()
                tipos_de_situacoes.update(ordem=F("ordem") - 1)

            if ordem_anterior > ordem_nova:
                tipos_de_situacoes = models.Tipo_de_situacao.objects.filter(
                    ordem__range=(ordem_nova, ordem_anterior)
                ).all()
                tipos_de_situacoes.update(ordem=F("ordem") + 1)

            tipo_de_situacao.ordem = ordem_nova
            tipo_de_situacao.save()

            queryset = models.Tipo_de_situacao.objects.all()
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class serve_protected_document(generics.RetrieveAPIView):
    model = models.Documento
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = serializers.DocumentoSerializer

    def get(self, request, *args, **kwargs):
        folder = self.kwargs.get("folder")
        file = self.kwargs.get("file")
        document = get_object_or_404(models.Documento, arquivo="documentos/" + folder + "/" + file)
        path, file_name = os.path.split(file)
        if os.path.isfile(document.arquivo.path):
            response = FileResponse(
                document.arquivo,
            )
            response["Content-Disposition"] = "attachment; filename=" + file_name
            return response
        raise Http404


class downloadDocumentsFromProcesso(generics.RetrieveAPIView):
    model = models.Documento
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = serializers.DocumentoSerializer

    def get(self, request, *args, **kwargs):
        processo_id = self.request.query_params.get("processo_id", None)
        if processo_id:
            filePaths = []
            queryset = models.Documento.objects.distinct().filter(processo__id=processo_id)
            for documento in queryset.all():
                filePaths.append(documento.arquivo.path)
                print(os.path.basename(documento.arquivo.path))
                print(os.path.splitext(documento.arquivo.path)[1])
            zip_subdir = "documentos"
            zip_filename = "%s.zip" % zip_subdir
            # with ZipFile('sample2.zip', 'w') as zipObj2:
            #     # Add multiple files to the zip
            #     zipObj2.write('sample_file.csv')
            s = BytesIO()
            zf = zipfile.ZipFile(s, "w")
            for fpath in filePaths:
                fdir, fname = os.path.split(fpath)
                zip_path = os.path.join(zip_subdir, fname)
                zf.write(fpath, zip_path)
            zf.close()

            response = HttpResponse(s.getvalue(), content_type=("application/zip"))
            response["Content-Disposition"] = "attachment; filename=%s" % zip_filename
            return response

        return Response(status=status.HTTP_400_BAD_REQUEST)


class downloadTodosProcessos(generics.RetrieveAPIView):
    model = models.Processo
    permission_classes = [permissions.IsAdminUser]
    serializer_class = serializers.ProcessoSerializer

    def get(self, request, *args, **kwargs):

        output = BytesIO()

        book = xlsxwriter.Workbook(output)
        sheet = book.add_worksheet("Processos")

        processo_index = 1
        sheet.write(processo_index, 1, "Data de Criação")
        sheet.write(processo_index, 2, "Processo")
        sheet.write(processo_index, 3, "Auto de Infração")
        sheet.write(processo_index, 4, "Reclamante")
        sheet.write(processo_index, 5, "Reclamada")
        sheet.write(processo_index, 6, "CPF/CNPJ")
        sheet.write(processo_index, 7, "Última Situação")
        sheet.write(processo_index, 8, "Data da última situação")
        sheet.write(processo_index, 9, "Ficha de Atendimento")
        processo_index += 1
        for processo in models.Processo.objects.all():
            ultima_situacao = processo.ultima_situacao()
            sheet.write(processo_index, 1, processo.criado_em.strftime("%d/%m/%Y %H:%M:%S"))
            sheet.write(processo_index, 2, processo.identificacao)
            sheet.write(processo_index, 3, processo.auto_infracao)
            sheet.write(processo_index, 4, processo.reclamante)
            sheet.write(processo_index, 5, processo.reclamada)
            sheet.write(processo_index, 6, processo.cpf_cnpj)
            if ultima_situacao:
                sheet.write(processo_index, 7, ultima_situacao.tipo_de_situacao.nome)
                sheet.write(processo_index, 8, ultima_situacao.data.strftime("%d/%m/%Y %H:%M:%S"))
            sheet.write(processo_index, 9, processo.ficha_de_atendimento)
            processo_index += 1
        book.close()

        # construct response
        output.seek(0)
        response = HttpResponse(
            output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=test.xlsx"

        return response


class exportarProcessos(views.APIView):
    model = models.Processo
    permission_classes = [permissions.IsAdminUser]
    serializer_class = serializers.ProcessoSerializer
    parser_classes = (MultiPartParser, FormParser)

    def put(self, request):
        print("Iniciando exportação:")
        file_obj = request.FILES["planilha"]
        wb = load_workbook(file_obj)
        sheet = wb.active
        tipo_de_situacao_ordem = models.Tipo_de_situacao.objects.all().count() + 1
        for row_count in range(2, sheet.max_row):
            processo = models.Processo()
            criado_em = sheet.cell(row=row_count, column=1).internal_value
            if isinstance(criado_em, datetime):
                processo.criado_em = timezone.make_aware(criado_em)
            identificacao = sheet.cell(row=row_count, column=2).internal_value
            if identificacao:
                processo.identificacao = str(identificacao)
            auto_infracao = sheet.cell(row=row_count, column=3).internal_value
            if auto_infracao:
                processo.auto_infracao = str(auto_infracao)
            reclamante = sheet.cell(row=row_count, column=4).internal_value
            if reclamante:
                processo.reclamante = str(reclamante)
            reclamada = sheet.cell(row=row_count, column=5).internal_value
            if reclamada:
                processo.reclamada = str(reclamada)
            cpf_cnpj = sheet.cell(row=row_count, column=6).internal_value
            if cpf_cnpj:
                cpf_cnpj = str(cpf_cnpj)
                # "999.999.999-99" 11
                mask_cpf = "%s%s%s.%s%s%s.%s%s%s-%s%s"
                # "99.999.999/9999-99" 14
                mask_cnpj = "%s%s.%s%s%s.%s%s%s/%s%s%s%s-%s%s"
                cpj_cnpj_onlynumbers = "".join(filter(str.isdigit, cpf_cnpj))
                if len(cpj_cnpj_onlynumbers) == 11:
                    cpf_cnpj = mask_cpf % tuple(cpj_cnpj_onlynumbers)
                elif len(cpj_cnpj_onlynumbers) == 14:
                    cpf_cnpj = mask_cnpj % tuple(cpj_cnpj_onlynumbers)
                processo.cpf_cnpj = cpf_cnpj

            ficha_de_atendimento = str(sheet.cell(row=row_count, column=9).internal_value)
            if ficha_de_atendimento:
                processo.ficha_de_atendimento = str(ficha_de_atendimento)

            if processo.identificacao and processo.identificacao != "None":
                print("Processo: " + processo.identificacao + " | Linha: " + str(row_count))
                processo.save()
                if processo.id:

                    tipo_de_situacao_nome = sheet.cell(row=row_count, column=7).internal_value
                    if tipo_de_situacao_nome:
                        tipo_de_situacao_nome = str(tipo_de_situacao_nome)
                        situacao_data = sheet.cell(row=row_count, column=8).internal_value

                        tipo_de_situacao = models.Tipo_de_situacao.objects.filter(
                            nome__iexact=tipo_de_situacao_nome
                        ).first()
                        if not tipo_de_situacao:
                            tipo_de_situacao = models.Tipo_de_situacao.objects.create(
                                nome=tipo_de_situacao_nome, ordem=tipo_de_situacao_ordem
                            )
                            tipo_de_situacao_ordem += 1
                        if tipo_de_situacao:
                            situacao = models.Situcacao()
                            situacao.processo = processo
                            situacao.tipo_de_situacao = tipo_de_situacao
                            if isinstance(situacao_data, datetime):
                                situacao.data = timezone.make_aware(situacao_data)
                            situacao.save()

        # do some stuff with uploaded file
        return Response(status=204)
